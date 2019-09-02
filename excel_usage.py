#repetition controller
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook

kitap = load_workbook("Kitap.xlsx")#Open xlsx file
sayfa = kitap.active#and active
bosliste=[]

for i in range(1,10):
    kd = sayfa.cell(i,1)#Cell values (1,1)
    if(kd.value in bosliste):
        print("Bu veri daha önce eklendi = Kolon : 1 - Satır : {}".format(i))
    elif(kd.value is None):
        print("Burası Boş = Kolon : 1 - Satır : {}".format(i))
    else:
        bosliste.append(kd.value)

kitap.close()